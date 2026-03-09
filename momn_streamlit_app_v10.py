# ============================================================
#  Momentum Ranking App  — v10.0
#  Multi-API data source: YFinance | Upstox | Zerodha
#
#  Changelog v10.0:
#    - Added "Data Source" dropdown (YFinance / Upstox / Zerodha)
#    - Extracted all data-fetching into data_service.py
#    - YFinance path unchanged from v9.1
#    - Upstox & Zerodha: Phase-1 mock data; Phase-2 hooks in data_service.py
#    - Consistent (close, high, volume) output contract across all APIs
#    - All downstream momentum calculations untouched
#
#  V6.0: added login page
#  V7.0: added Reason for exit after portfolio rebalancing
#  V8.0: added failed stock download list via volm_cr analysis
#  V9.1: removed cond8 "1M ROC / 12M ROC ratio < 50%"
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import random
from datetime import datetime
from dateutil.relativedelta import relativedelta
import time
import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from json.decoder import JSONDecodeError

# ── NEW: import unified data-fetching service ─────────────────
from data_service import fetch_data

# ─────────────────────────────────────────────────────────────
# Hard-coded app credentials
# ─────────────────────────────────────────────────────────────
USERNAME = "prayan"
PASSWORD = "prayan"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False


# ─────────────────────────────────────────────────────────────
# LOGIN PAGE
# ─────────────────────────────────────────────────────────────
def login():
    st.title("Login")
    with st.form(key="login_form", clear_on_submit=True):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit_button = st.form_submit_button(label="Login")
        if submit_button:
            if username == USERNAME and password == PASSWORD:
                st.session_state.logged_in = True
                st.success("Logged in successfully!")
                st.rerun()
            else:
                st.error("Invalid username or password")


# ─────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────
def app_content():

    # ── Helper / metric functions (unchanged from v9.1) ──────────
    @st.cache_data(ttl=0)
    def getMedianVolume(data):
        return round(data.median(), 0)

    def getDailyReturns(data):
        return data.pct_change(fill_method='ffill')

    def getMaskDailyChange(data):
        m1 = getDailyReturns(data).eq(np.inf)
        m2 = getDailyReturns(data).eq(-np.inf)
        return (getDailyReturns(data)
                .mask(m1, df[~m1].max(), axis=1)
                .mask(m2, df[~m2].min(), axis=1)
                .bfill(axis=1))

    def getStdev(data):
        return np.std(getMaskDailyChange(data) * 100)

    def getStdRatio(data, data1):
        return getStdev(data) / getStdev(data1) * 100

    def getAbsReturns(data):
        return round((data.iloc[-1] / data.iloc[0] - 1) * 100, 2)

    def getVolatility(data):
        return round(np.std(data) * np.sqrt(252) * 100, 2)

    def getMonthlyPrices(data):
        grps = data.groupby([data.index.year, data.index.month])
        monthly = pd.DataFrame()
        for k in grps:
            monthly = pd.concat([monthly, k[1].tail(1)])
        return monthly

    def getMonthlyReturns(data):
        return data.pct_change()

    def getSharpe(data):
        return round(np.sqrt(252) * data.mean() / data.std(), 2)

    def getSortino(data):
        return np.sqrt(252) * data.mean() / data[data < 0].std()

    def getMaxDrawdown(data):
        cummRet = (data + 1).cumprod()
        peak = cummRet.expanding(min_periods=1).max()
        return ((cummRet / peak) - 1).min()

    def getCalmar(data):
        return data.mean() * 252 / abs(getMaxDrawdown(data))

    def getAbsMomentumVolAdjusted(absReturn, volatility):
        return absReturn / volatility

    def getNMonthRoC(data, N):
        return round((data.iloc[-1] / data.iloc[-1 - N] - 1) * 100, 2)

    def getNWeekRoC(data, N):
        return round((data.iloc[-1] / data.iloc[-1 - N] - 1) * 100, 2)

    def getFIP(data):
        retPos = np.sum(data.pct_change()[1:] > 0)
        retNeg = np.sum(data.pct_change()[1:] < 0)
        return retPos - retNeg

    def getSharpeRoC(roc, volatility):
        return round(roc / volatility, 2)

    def getBeta(dfNifty, data12M):
        dailyReturns = getDailyReturns(pd.concat([dfNifty, data12M], axis=1))[1:]
        var = np.var(dailyReturns['Nifty'])
        cov = dailyReturns.cov()
        return [round(cov.loc[k, 'Nifty'] / var, 2) for k in cov.columns[1:]]

    def calculate_z_score(data):
        mean, std = data.mean(), data.std()
        return ((data - mean) / std).round(2)

    # ─────────────────────────────────────────────────────────────
    # UI — Title + Controls
    # ─────────────────────────────────────────────────────────────
    st.title("Momentum Ranking App")

    import warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)

    # ── Ranking method ────────────────────────────────────────────
    ranking_options = {
        "AvgZScore 12M/6M/3M":       "avgZScore12_6_3",
        "AvgZScore 12M/9M/6M/3M":    "avgZScore12_9_6_3",
        "AvgSharpe 12M/6M/3M":       "avgSharpe12_6_3",
        "AvgSharpe 9M/6M/3M":        "avgSharpe9_6_3",
        "AvgSharpe 12M/9M/6M/3M":    "avg_All",
        "Sharpe12M":                  "sharpe12M",
        "Sharpe3M":                   "sharpe3M",
    }
    ranking_method_display = st.selectbox(
        "Select Ranking Method",
        options=list(ranking_options.keys()),
        index=0
    )
    ranking_method = ranking_options[ranking_method_display]

    # ── Universe ──────────────────────────────────────────────────
    universe = ['Nifty50', 'Nifty100', 'Nifty200', 'Nifty250', 'Nifty500', 'N750', 'AllNSE']
    U = st.selectbox('Select Universe:', universe, index=6)

    # ── ✨ NEW: Data Source dropdown ──────────────────────────────
    #    Phase-1: YFinance = live data; Upstox & Zerodha = mock data
    #    Phase-2: set real credentials in data_service.py → UPSTOX_CONFIG / ZERODHA_CONFIG
    API_OPTIONS = ["YFinance", "Upstox", "Zerodha"]
    api_source = st.selectbox(
        "Select Data Source",
        options=API_OPTIONS,
        index=0,
        help=(
            "YFinance: live data (default).\n"
            "Upstox / Zerodha: mock data in Phase-1. "
            "Add real credentials in data_service.py to activate."
        )
    )

    # Show a Phase-1 notice for non-YFinance sources
    if api_source in ("Upstox", "Zerodha"):
        st.info(
            f"ℹ️ **{api_source}** is running in **mock/demo mode** (Phase-1). "
            f"To use real data, add your API credentials in `data_service.py` "
            f"under `{api_source.upper()}_CONFIG` and uncomment the Phase-2 block.",
            icon="🔧"
        )

    # ── Lookback date ─────────────────────────────────────────────
    selected_date = st.date_input("Select Lookback Date", datetime.today())
    dt2 = datetime.strptime(str(selected_date), "%Y-%m-%d").strftime('%Y-%m-%d')

    dates = {
        'startDate': datetime.strptime('2000-01-01', '%Y-%m-%d'),
        'endDate':   datetime.strptime(dt2, '%Y-%m-%d'),
        'date12M':   datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=12),
        'date9M':    datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=9),
        'date6M':    datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=6),
        'date3M':    datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=3),
        'date1M':    datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=1),
    }

    st.write("##### Date Range:")
    st.write(f"Start Date: **{dates['startDate'].strftime('%d-%m-%Y')}**")
    st.write(f"End Date: **{dates['endDate'].strftime('%d-%m-%Y')}**")

    # ─────────────────────────────────────────────────────────────
    # Load symbol list from CSV (unchanged)
    # ─────────────────────────────────────────────────────────────
    if U == 'N750':
        file_path = 'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/ind_niftytotalmarket_list.csv'
    elif U == 'AllNSE':
        file_path = 'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/NSE_EQ_ALL.csv'
    else:
        file_path = f'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/ind_{U.lower()}list.csv'

    df = pd.read_csv(file_path)
    df['Yahoo_Symbol'] = df.Symbol + '.NS'
    df = df.set_index('Yahoo_Symbol')
    symbol = list(df.index)

    CHUNK = 15 if U == "AllNSE" else 50
    st.write(f"Chunk size set to **{CHUNK}** for universe: **{U}** | Data source: **{api_source}**")

    # ─────────────────────────────────────────────────────────────
    # START DOWNLOAD BUTTON
    # ─────────────────────────────────────────────────────────────
    start_button = st.button(f"▶ Start Data Download ({api_source})")

    if start_button:
        progress_bar = st.progress(0)
        status_text  = st.empty()

        # ── ✨ UNIFIED DATA FETCH — single call regardless of API ────
        #    Hook: swap api_source to change data provider.
        #    All three return identical (close, high, volume, failed_symbols).
        with st.spinner(f"Fetching data from {api_source}…"):
            close, high, volume, failed_symbols = fetch_data(
                api_source  = api_source,
                symbols     = symbol,
                start_date  = dates['startDate'],
                end_date    = dates['endDate'],
                chunk_size  = CHUNK,
                progress_bar= progress_bar,
                status_text = status_text,
            )

        # ── Failed symbol report ──────────────────────────────────
        volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
        median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
        failed_blank    = median_volume[median_volume.isna()].index.tolist()
        failed_blank    = [t.replace('.NS', '') for t in failed_blank]

        if failed_blank:
            st.warning("The following stocks failed to download (blank volume data):")
            failed_table = pd.DataFrame({
                'S.No.': range(1, len(failed_blank) + 1),
                'Failed Stocks': failed_blank
            }).set_index('S.No.')
            st.dataframe(failed_table)
        else:
            st.success("All stocks downloaded successfully!")

        st.write("All data download attempts complete.")

        # ─────────────────────────────────────────────────────────
        # NEXT REBALANCE DATE (unchanged)
        # ─────────────────────────────────────────────────────────
        def get_next_rebalance_date(current_date):
            current_date = pd.Timestamp(current_date)
            rebalance_months = [3, 9]
            rebalance_dates = []
            for month in rebalance_months:
                last_day = pd.Timestamp(current_date.year, month, 1) + pd.offsets.MonthEnd(0)
                while last_day.weekday() >= 5:
                    last_day -= dt.timedelta(days=1)
                rebalance_dates.append(last_day)
            for date in sorted(rebalance_dates):
                if date > current_date:
                    return date.date()
            next_year = []
            for month in rebalance_months:
                last_day = pd.Timestamp(current_date.year + 1, month, 1) + pd.offsets.MonthEnd(0)
                while last_day.weekday() >= 5:
                    last_day -= dt.timedelta(days=1)
                next_year.append(last_day)
            return sorted(next_year)[0].date()

        current_date = dt.date.today()
        next_rebalance_date = get_next_rebalance_date(current_date)
        formatted_rebalance_date = next_rebalance_date.strftime("%d-%m-%Y")

        filters = [
            "Avg. Daily Volume greater than 1 crore (volm_cr > 1)",
            "Closing price above 200-day moving average (Close > dma200d)",
            "12-month rate of change (ROC) greater than 6.5% (roc12M > 6.5)",
            "Number of circuit hits in a year less than 20 (circuit < 20)",
            "Stock is within 25% of its all-time high (AWAY_ATH > -25)",
            "12-month return less than 10x (roc12M < 1000)",
            "Closing price above ₹30 (Close > 30)",
            "No more than 10 circuits of 5% in the last 3 months (circuit5 <= 10)",
        ]

        with st.sidebar:
            st.header("Menu")
            st.info(f"Index Rebalance Date: **{formatted_rebalance_date}**.")
            with st.expander("Applied Filters", expanded=False):
                st.write("The following conditions are applied:")
                for i, f in enumerate(filters, start=1):
                    st.write(f"{i}. {f}")

        # ─────────────────────────────────────────────────────────
        # SLICE DATA TO DATE RANGES (unchanged)
        # ─────────────────────────────────────────────────────────
        data20Y   = close.loc[:dates['endDate']].copy()
        volume20Y = volume.loc[:dates['endDate']].copy()
        high20Y   = high.loc[:dates['endDate']].copy()

        data12M   = data20Y.loc[dates['date12M']:].copy()
        data9M    = data20Y.loc[dates['date9M']:].copy()
        data6M    = data20Y.loc[dates['date6M']:].copy()
        data3M    = data20Y.loc[dates['date3M']:].copy()
        data1M    = data20Y.loc[dates['date1M']:].copy()
        volume12M = volume20Y.loc[dates['date12M']:].copy()

        # ─────────────────────────────────────────────────────────
        # CALCULATE METRICS (unchanged from v9.1)
        # ─────────────────────────────────────────────────────────
        dfStats = pd.DataFrame(index=symbol)
        dfStats['Close']   = round(data12M.iloc[-1], 2)
        data12M_Temp = data12M.fillna(0)
        dfStats['dma200d'] = round(data12M_Temp.rolling(window=200).mean().iloc[-1], 2)

        dfStats['roc12M'] = getAbsReturns(data12M)
        dfStats['roc9M']  = getAbsReturns(data9M)
        dfStats['roc6M']  = getAbsReturns(data6M)
        dfStats['roc3M']  = getAbsReturns(data3M)
        dfStats['roc1M']  = getAbsReturns(data1M)

        dfStats['vol12M'] = getVolatility(getDailyReturns(data12M))
        dfStats['vol9M']  = getVolatility(getDailyReturns(data9M))
        dfStats['vol6M']  = getVolatility(getDailyReturns(data6M))
        dfStats['vol3M']  = getVolatility(getDailyReturns(data3M))

        dfStats['sharpe12M'] = getSharpeRoC(dfStats['roc12M'], dfStats['vol12M'])
        dfStats['sharpe9M']  = getSharpeRoC(dfStats['roc9M'],  dfStats['vol9M'])
        dfStats['sharpe6M']  = getSharpeRoC(dfStats['roc6M'],  dfStats['vol6M'])
        dfStats['sharpe3M']  = getSharpeRoC(dfStats['roc3M'],  dfStats['vol3M'])

        dfStats['z_score12M'] = calculate_z_score(dfStats['sharpe12M'])
        dfStats['z_score9M']  = calculate_z_score(dfStats['sharpe9M'])
        dfStats['z_score6M']  = calculate_z_score(dfStats['sharpe6M'])
        dfStats['z_score3M']  = calculate_z_score(dfStats['sharpe3M'])

        for col in ['sharpe12M', 'sharpe9M', 'sharpe6M', 'sharpe3M',
                    'z_score12M', 'z_score9M', 'z_score6M', 'z_score3M']:
            dfStats[col] = dfStats[col].replace([np.inf, -np.inf], np.nan).fillna(0)

        # ── Ranking score columns ─────────────────────────────────
        if ranking_method == "avgSharpe12_6_3":
            dfStats['avgSharpe12_6_3'] = dfStats[["sharpe12M", "sharpe6M", "sharpe3M"]].mean(axis=1).round(2)
        elif ranking_method == "avg_All":
            dfStats['avg_All'] = dfStats[["sharpe12M", "sharpe9M", "sharpe6M", "sharpe3M"]].mean(axis=1).round(2)
        elif ranking_method == "avgSharpe9_6_3":
            dfStats['avgSharpe9_6_3'] = dfStats[["sharpe9M", "sharpe6M", "sharpe3M"]].mean(axis=1).round(2)
        elif ranking_method == "avgZScore12_6_3":
            dfStats['avgZScore12_6_3'] = dfStats[['z_score12M', 'z_score6M', 'z_score3M']].mean(axis=1).round(2)
        elif ranking_method == "avgZScore12_9_6_3":
            dfStats['avgZScore12_9_6_3'] = dfStats[['z_score12M', 'z_score9M', 'z_score6M', 'z_score3M']].mean(axis=1).round(2)

        dfStats['volm_cr'] = (getMedianVolume(volume12M) / 1e7).round(2)
        dfStats['ATH']     = round(high20Y.max(), 2)
        dfStats['AWAY_ATH']= round((dfStats['Close'] / dfStats['ATH'] - 1) * 100, 2)

        dataDaily_pct = round(getDailyReturns(data12M) * 100, 2)
        dfStats['circuit'] = (
            (dataDaily_pct ==  4.99).sum() + (dataDaily_pct ==  5.00).sum() +
            (dataDaily_pct ==  9.99).sum() + (dataDaily_pct == 10.00).sum() +
            (dataDaily_pct == 19.99).sum() + (dataDaily_pct == 20.00).sum() +
            (dataDaily_pct == -4.99).sum() + (dataDaily_pct == -5.00).sum() +
            (dataDaily_pct == -9.99).sum() + (dataDaily_pct == -10.00).sum() +
            (dataDaily_pct == -19.99).sum() + (dataDaily_pct == -20.00).sum()
        )

        dataDaily_pct5 = round(getDailyReturns(data3M) * 100, 2)
        dfStats['circuit5'] = (
            (dataDaily_pct5 ==  4.99).sum() + (dataDaily_pct5 ==  5.00).sum() +
            (dataDaily_pct5 == -4.99).sum() + (dataDaily_pct5 == -5.00).sum()
        )

        # ── Ticker cleanup ────────────────────────────────────────
        dfStats = dfStats.reset_index().rename(columns={'index': 'Ticker'})
        dfStats['Ticker'] = dfStats['Ticker'].astype(str).str.replace('.NS', '', regex=False)

        for col in ['avgSharpe12_6_3', 'avg_All', 'avgSharpe9_6_3',
                    'avgZScore12_6_3', 'avgZScore12_9_6_3']:
            if col in dfStats.columns:
                dfStats[col] = dfStats[col].replace([np.inf, -np.inf], np.nan).fillna(0)

        # ── Sort & rank ───────────────────────────────────────────
        if ranking_method in ["avg_All", "sharpe12M"]:
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc12M'], ascending=[False, False])
        elif ranking_method in ["avgSharpe12_6_3", "sharpe3M"]:
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc3M'],  ascending=[False, False])
        elif ranking_method == "avgSharpe9_6_3":
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc6M'],  ascending=[False, False])
        elif ranking_method == "avgZScore12_6_3":
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc3M'],  ascending=[False, False])
        elif ranking_method == "avgZScore12_9_6_3":
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc6M'],  ascending=[False, False])

        dfStats['Rank'] = range(1, len(dfStats) + 1)
        dfStats = dfStats.set_index('Rank')

        # ─────────────────────────────────────────────────────────
        # DISPLAY UNFILTERED DATA
        # ─────────────────────────────────────────────────────────
        st.info("Unfiltered Data:")
        st.write(dfStats)

        # ─────────────────────────────────────────────────────────
        # APPLY FILTERS (unchanged)
        # ─────────────────────────────────────────────────────────
        cond1  = dfStats['volm_cr']   > 1
        cond3  = dfStats['Close']     > dfStats['dma200d']
        cond4  = dfStats['roc12M']    > 6.5
        cond5  = dfStats['circuit']   < 20
        cond6  = dfStats['AWAY_ATH']  > -25
        cond7  = dfStats['roc12M']    < 1000
        cond9  = dfStats['Close']     > 30
        cond10 = dfStats['circuit5']  <= 10

        dfStats['final_momentum'] = cond1 & cond3 & cond4 & cond5 & cond6 & cond7 & cond9 & cond10

        filtered = dfStats[dfStats['final_momentum']].sort_values('Rank', ascending=True)

        st.info("Filtered Data:")
        st.write(filtered)

        # ─────────────────────────────────────────────────────────
        # EXCEL EXPORT (unchanged from v9.1)
        # ─────────────────────────────────────────────────────────
        def format_excel(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border    = thin
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = 'A2'
            hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.fill      = hdr_fill
                c.font      = hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

            no_fill   = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
            bold_font = Font(bold=True)
            headers   = [c.value for c in ws[1]]

            def ci(name):
                return headers.index(name) + 1 if name in headers else None

            rank_threshold = 100 if U == 'AllNSE' else 75
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            idx = {k: ci(k) for k in ['volm_cr','Close','dma200d','AWAY_ATH','roc12M',
                                       'circuit','roc1M','circuit5','Ticker','Rank']}

            for row in range(2, ws.max_row + 1):
                failed = False
                def v(col): return ws.cell(row=row, column=col).value if col else None
                def mark(col):
                    nonlocal failed
                    ws.cell(row=row, column=col).fill = no_fill
                    ws.cell(row=row, column=col).font = bold_font
                    failed = True

                if (vol := v(idx['volm_cr'])) is not None and vol < 1:            mark(idx['volm_cr'])
                cl = v(idx['Close']); dm = v(idx['dma200d'])
                if cl is not None and dm is not None and cl <= dm:                 mark(idx['Close'])
                if (aa := v(idx['AWAY_ATH'])) is not None and aa <= -25:          mark(idx['AWAY_ATH'])
                roc = v(idx['roc12M'])
                if roc is not None and roc <= 6.5:                                 mark(idx['roc12M'])
                if (ci_ := v(idx['circuit'])) is not None and ci_ >= 20:          mark(idx['circuit'])
                if cl is not None and cl <= 30:                                    mark(idx['Close'])
                if (c5 := v(idx['circuit5'])) is not None and c5 > 10:            mark(idx['circuit5'])
                if roc is not None and roc > 1000:                                 mark(idx['roc12M'])
                if failed and idx['Ticker']:
                    ws.cell(row=row, column=idx['Ticker']).fill = no_fill
                if idx['Rank'] and (rk := v(idx['Rank'])) is not None and rk <= rank_threshold:
                    ws.cell(row=row, column=idx['Rank']).fill = green_fill

            # ATH column round
            ath_col = ci('ATH')
            if ath_col:
                for r in range(2, ws.max_row + 1):
                    c = ws.cell(row=r, column=ath_col)
                    if isinstance(c.value, (int, float)):
                        c.value = round(c.value)

            wb.save(file_name)

        def format_filtered_excel(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb["Filtered Stocks"]
            thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"),  bottom=Side(style="thin"))
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border    = thin
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = 'A2'
            hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

            # ATH round
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "ATH":
                    for r in range(2, ws.max_row + 1):
                        c = ws.cell(row=r, column=col)
                        if isinstance(c.value, (int, float)):
                            c.value = round(c.value)
                    break

            # AWAY_ATH % suffix
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "AWAY_ATH":
                    for r in range(2, ws.max_row + 1):
                        c = ws.cell(row=r, column=col)
                        if isinstance(c.value, (int, float)):
                            c.value = f"{c.value}%"
                    break

            # Rank highlight
            rank_threshold = 100 if U == 'AllNSE' else 75
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Rank":
                    rank_75_count = 0
                    for r in range(2, ws.max_row + 1):
                        c = ws.cell(row=r, column=col)
                        if isinstance(c.value, (int, float)) and c.value <= rank_threshold:
                            c.fill = green_fill
                            rank_75_count += 1
                    total_filtered = ws.max_row - 1
                    ws.append([])
                    ws.append(["Summary"])
                    summary_start = ws.max_row
                    ws.append([f"Total Filtered Stocks: {total_filtered}"])
                    ws.append([f"Number of Stocks within {rank_threshold} Rank: {rank_75_count}"])
                    for r in ws.iter_rows(min_row=summary_start, max_row=ws.max_row, min_col=1, max_col=1):
                        for cell in r:
                            cell.font = Font(bold=True)
                    break

            wb.save(file_name)

        excel_file = f"{selected_date.strftime('%Y-%m-%d')}_{U}_{ranking_method}_{api_source}_lookback.xlsx"
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            dfStats.to_excel(writer,  sheet_name="Unfiltered Stocks", index=True)
            filtered.to_excel(writer, sheet_name="Filtered Stocks",   index=True)

        format_excel(excel_file)
        format_filtered_excel(excel_file)

        st.download_button(
            label     = "Download Stock Data as Excel",
            data      = open(excel_file, "rb").read(),
            file_name = excel_file,
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ─────────────────────────────────────────────────────────
        # PORTFOLIO REBALANCING (unchanged from v9.1)
        # ─────────────────────────────────────────────────────────
        filtered = dfStats[dfStats['final_momentum']].sort_values('Rank', ascending=True)
        rank_threshold   = 100 if U == 'AllNSE' else 75
        top_rank_tickers = filtered[filtered.index <= rank_threshold]['Ticker']

        portfolio_url = (
            "https://docs.google.com/spreadsheets/d/e/"
            "2PACX-1vS4HDgiell4n1kd08OnlzOQobfPzeDtVyWJ8gETFlYbz27qhOmfqKZOoIXZItRQEq5ANATYIcZJm0gk"
            "/pub?output=csv"
        )

        with st.spinner("Portfolio Rebalancing… Please wait…"):
            portfolio_data = pd.read_csv(portfolio_url)

            if 'Current Portfolio' not in portfolio_data.columns:
                st.error("Column 'Current Portfolio' not found in the portfolio data.")
            else:
                current_portfolio_tickers = portfolio_data['Current Portfolio']
                entry_stocks = top_rank_tickers[~top_rank_tickers.isin(current_portfolio_tickers)]
                exit_stocks  = current_portfolio_tickers[~current_portfolio_tickers.isin(top_rank_tickers)]

                st.info("Portfolio Rebalancing:")
                num_sells   = len(exit_stocks)
                entry_stocks = entry_stocks.head(num_sells)

                if len(entry_stocks) < num_sells:
                    entry_stocks = pd.concat([
                        entry_stocks,
                        pd.Series([None] * (num_sells - len(entry_stocks)))
                    ])

                reasons_for_exit = []
                for ticker in exit_stocks:
                    if pd.isna(ticker) or ticker == "":
                        reasons_for_exit.append("")
                        continue
                    reasons    = []
                    stock_data = dfStats[dfStats['Ticker'] == ticker]
                    if len(stock_data) > 0:
                        if stock_data.index[0] > rank_threshold:
                            reasons.append(f"Rank > {rank_threshold}")
                        if stock_data['volm_cr'].values[0] <= 1:
                            reasons.append("Volume <= 1 crore")
                        if stock_data['Close'].values[0] <= stock_data['dma200d'].values[0]:
                            reasons.append("Close <= 200-day DMA")
                        if stock_data['roc12M'].values[0] <= 6.5:
                            reasons.append("12M ROC <= 6.5%")
                        if stock_data['circuit'].values[0] >= 20:
                            reasons.append("Circuit hits >= 20")
                        if stock_data['AWAY_ATH'].values[0] <= -25:
                            reasons.append("Away from ATH <= -25%")
                        if stock_data['roc12M'].values[0] >= 1000:
                            reasons.append("12M ROC >= 1000%")
                        if stock_data['Close'].values[0] <= 30:
                            reasons.append("Close <= 30")
                        if stock_data['circuit5'].values[0] > 10:
                            reasons.append("5% Circuit hits > 10")
                    else:
                        reasons.append("Stock not in selected universe")
                    reasons_for_exit.append(", ".join(reasons) if reasons else "")

                reasons_for_exit.extend([""] * (len(entry_stocks) - len(reasons_for_exit)))

                rebalance_table = pd.DataFrame({
                    'S.No.':          range(1, num_sells + 1),
                    'Sell Stocks':    exit_stocks.tolist(),
                    'Buy Stocks':     entry_stocks.tolist(),
                    'Reason for Exit': reasons_for_exit,
                })
                rebalance_table = rebalance_table[
                    ~(rebalance_table['Sell Stocks'].isna() & rebalance_table['Buy Stocks'].isna())
                ]
                rebalance_table.set_index('S.No.', inplace=True)
                st.dataframe(rebalance_table)

        st.success("Portfolio Rebalancing completed!")


# ─────────────────────────────────────────────────────────────
# ENTRYPOINT
# ─────────────────────────────────────────────────────────────
if not st.session_state.logged_in:
    login()
else:
    app_content()
